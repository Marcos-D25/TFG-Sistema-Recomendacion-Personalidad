import os
import pandas as pd
import numpy as np
import joblib
from openpyxl import load_workbook
import optuna
from sklearn.metrics import classification_report, confusion_matrix, precision_recall_fscore_support, accuracy_score

from balanceador import Balanceador, BalanceadorSMOTE, BalanceadorBorderlineSMOTE, BalanceadorADASYN, BalanceadorENN, BalanceadorAKNN
from procesador import Preprocesador
from clasificador import Clasificador, XGB, LR, MLPC,LSVC, DTC, KNC

class Pipeline:
    def __init__(self, nombre_modelo, balanceador:Balanceador=None):
        self.nombre_modelo = nombre_modelo
        self.nombre_balanceador = balanceador.__str__() if balanceador else None
        self.balanceador = balanceador

    def entreno_clasificador(self, buscar_hiper:bool=False,parametros:dict=None, guardar:bool=True, nomExcel:str="resultados.xlsx") -> None:
        '''
        Funcion que entrena el clasificador elegido.

        :param buscar_hiper: Bool que decide si buscar la mejor combinacion de hiperparametros antes de realizar el entreno. Si True se ignorarán los parametros recibidos en la función
        :param parametros: Diccionario que contiene los parametros especificos para el entreno. En caso de no indicar ninguno, se usarán los parametros predefinidos de la clase
        :param guardar: Bool que indica si se guardan los modelos en una carpeta local con el nombre del clasificador
        :param nomExcel: Nombre del archivo excel en el que guardar los resultados de los modelos entrenados
        :return: None
        '''

        if buscar_hiper:
            parametros = self.clasificador.busqueda_hiperparametros()
        elif parametros is None: 
            parametros = self.clasificador.getParametros()
        
        self.clasificador.entrenar_modelo()

        self.modelos = self.clasificador.getModelos()
        
        if guardar:
            self.clasificador.guardar_modelo(f"modelos_{self.clasificador.__str__()}", self.nombre_modelo.replace("/","_"))

        self.guardar_resultados(nombre_Archivo=nomExcel, metodo_balanceo=self.nombre_balanceador, parametros_str=str(parametros), modelo_clasificacion=self.clasificador.__str__())

    def obtener_metricas(self, modelo:Clasificador, df_test:pd.DataFrame, nombre_modelo:str) -> dict:
        '''
        Funcion que devuelve las metricas Precision, Recall, F1-Score, Support, Accuracy y Matriz Confusion del modelo entrenado

        :param modelo: Modelo clasificador entrenado
        :param df_test: Dataframe que contiene los datos para el proceso de test
        :param nombre_modelo: Nombre del modelo clasificador usado para el entreno
        :return: Diccionario con los datos de entreno 
        '''
        X_test = np.array(df_test["Embedding"].tolist(), dtype=np.float32)
        y_test = df_test["MBTI"].tolist()
        y_pred = modelo.predict(X_test)
        
        precision, recall, f1, support = precision_recall_fscore_support(y_test, y_pred, labels=[0, 1])
        acc = accuracy_score(y_test, y_pred)
        cm = confusion_matrix(y_test, y_pred)
        
        return {
            "Modelos": nombre_modelo,
            "Precision": f"0: {precision[0]:.2f}\n1: {precision[1]:.2f}",
            "Recall": f"0: {recall[0]:.2f}\n1: {recall[1]:.2f}",
            "F1-Score": f"0: {f1[0]:.2f}\n1: {f1[1]:.2f}",
            "Support": f"0: {support[0]}\n1: {support[1]}",
            "Accuracy": round(acc, 2),
            "Matriz Confusion": f"[[{cm[0][0]} {cm[0][1]}]\n [{cm[1][0]} {cm[1][1]}]]"
        }

    def guardar_resultados(self, nombre_Archivo:str="resultados.xlxs", carpeta:str="resultados", metodo_balanceo:str=None, parametros_str:dict=None, modelo_clasificacion=None):
        '''
        Funcion que guarda los resultados del entreno en el archivo excel correspondiente.

        :param nombre_Archivo: Nombre del archivo excel en el que se van a guardar los datos de entreno.
        :param carpeta: Nombre de la carpeta en la que se guardará el archivo con los resultados
        :param metodo_balanceo: Nombre del metodo de balanceo aplicado
        :param parametros_str: Diccionario que contiene los parametros que ha usado el modelo clasificador durante el entrenamiento
        :param modelo_clasificacion: Nombre del algoritmo de clasificacion usado
        :return: None
        '''

        print(f"[INFO] Exportando resultados a Excel (Pestaña: {metodo_balanceo})...")
    
        filas = []
        filas.append(self.obtener_metricas(self.modelos['E-I'], self.balanceador.test_EI, f"{modelo_clasificacion} E/I"))
        filas.append(self.obtener_metricas(self.modelos['S-N'], self.balanceador.test_SN, f"{modelo_clasificacion} S/N"))
        filas.append(self.obtener_metricas(self.modelos['T-F'], self.balanceador.test_TF, f"{modelo_clasificacion} T/F"))
        filas.append(self.obtener_metricas(self.modelos['J-P'], self.balanceador.test_JP, f"{modelo_clasificacion} J/P"))
        
        df_resultados = pd.DataFrame(filas)
        
        archivo_excel = os.path.join(carpeta, nombre_Archivo)
        
        modo = 'a' if os.path.exists(archivo_excel) else 'w'
        motor = 'openpyxl'
        fila_inicio = 0
        escribir_cabecera = True
        if modo == 'a':
            try:
                wb = load_workbook(archivo_excel)
                if metodo_balanceo in wb.sheetnames:
                    # La pestaña existe. Buscamos la última fila escrita y le sumamos 2 de margen
                    fila_inicio = wb[metodo_balanceo].max_row + 2
                    escribir_cabecera = False # Ya hay cabeceras arriba, no las repetimos
            except Exception as e:
                print(f"Aviso al leer el Excel: {e}")    


        with pd.ExcelWriter(archivo_excel, engine=motor, mode=modo, if_sheet_exists='overlay' if modo == 'a' else None) as writer:
            df_params = pd.DataFrame([["Parametros", parametros_str]])
            df_params.to_excel(writer, sheet_name=metodo_balanceo, startrow=fila_inicio, index=False, header=False)
            df_resultados.to_excel(writer, sheet_name=metodo_balanceo, startrow=fila_inicio + 2, index=False, header=escribir_cabecera)

        print(f"[INFO] Resultados guardados en {archivo_excel}")
    
    def ejecutar_pipeline_entreno(self, parametros:dict=None, buscar_hiperparametros:bool=False, algotitmo:str=None, nombre_archivo:str="resultados.xlsx"):
        '''
        Funcion que entrena al Clasificador y guarda los resultados en el archivo especificado.

        :param parametros: Parametros que se quieren usar sobre el Clasificador
        :param buscar_hiperparametros: Bool que indica si quiere que antes del entrenamiento se lleve acabo una busqueda de la mejor combinacion de hiperparametros (Llevará un rato)
        :param algoritmo: Nombre del algoritmo clasificador a usar [RL, XGBoost, LinearSVC, KNC, DTC, MLP]
        :param nombre_archivo: Nombre del archivo excel destino en el que guardar los resultados 
        '''

        print("[EJECUCION] Ejecutando pipeline completo ...")
        
        print("[EJECUCION] Dividiendo y balanceando dataset ...")
        self.balanceador.procesar_todos_ejes()

        self.clasificador:Clasificador = None

        match algotitmo:
            case "RL": 
                print(f"[EJECUCION] Entrenando modelo de Regresión Logística con {self.nombre_balanceador} ...")
                self.clasificador = LR(balanceador=self.balanceador)
            case "XGBoost":
                print(f"[EJECUCION] Entrenando modelo XGBoost con {self.nombre_balanceador} ...")
                self.clasificador = XGB(balanceador=self.balanceador)
            case "LinearSVC":
                print(f"[EJECUCION] Entrenando modelo LinearSVC con {self.nombre_balanceador} ...")
                self.clasificador = LSVC(balanceador=self.balanceador)
            case "KNC":
                print(f"[EJECUCION] Entrenando modelo KNeighborsClassifier con {self.nombre_balanceador}...")
                self.clasificador = KNC(balanceador=self.balanceador)
            case "DTC":
                print(f"[EJECUCION] Entrenando modelo DecisionTreeClassifier con {self.nombre_balanceador}...")
                self.clasificador = DTC(balanceador=self.balanceador)
            case "MLP":
                print(f"[EJECUCION] Entrenando modelo MultiLayerPerceptron con {self.nombre_balanceador}...")
                self.clasificador = MLPC(balanceador=self.balanceador)
            case _ : 
                print("[ERROR] Modelo de clasificación no reconocido")
                return None        
        
        self.entreno_clasificador(buscar_hiper=buscar_hiperparametros, parametros=parametros, nomExcel=nombre_archivo)

        print("[EJECUCION] Fin pipeline completo ...")

#Funcion que sirve para sacar las metricas de entrenamiento para un modelo de embedding concreto
def pipeline_modelo_entreno(modelo, preprocesar:bool = False,balancear=True, parametros:dict=None, buscar_hiperparametros:bool=False, carpeta_origen:str="dataset9K", modelos:list[str]=["FacebookAI/roberta-base"], carpeta_dest="datasetRB") -> None:
    '''
    Funcion que lleva acabo la ejecucion completa del proceso de entrenamiento. Cuenta con las fases de Preprocesamiento, Balanceo y entrenamiento del Clasificador

    :param modelo: Nombre del modelo encoder que se usa en el preprocesamiento. Tambien es el nombre del archivo excel donde se guardarán los resultados del entrenamiento
    :param preprocesar: Booleano que indica si se quiere preprocesar el dataset antes de entrar al proceso de entrenamiento. Si se cuenta con el dataset preprocesado, no haria falta volver a preprocesarlo de nuevo
    :param balancear: Booleano que indica si se quieren aplicar tecnicas de balanceo a los datasets
    :param parametros: Parametros que se quieren usar sobre el Clasificador
    :param buscar_hiperparametros: Bool que indica si quiere que antes del entrenamiento se lleve acabo una busqueda de la mejor combinacion de hiperparametros (Llevará un rato)
    :param carpeta_origen: Nombre de la carpeta del dataset original (Solo sirve si preprocesar = True)
    :param modelos: Lista de modelos encoders. Los nombres deben coincidir con el archivo local o modelo de HuggingFace. (Solo sirve si preprocesar = True)
    :param carpeta_dest: Nombre de la carpeta donde se guardaran los dataset preprocesados
    :return: None 
    '''
    print("[EJECUCION] Preprocesando dataset ...")
    if preprocesar:
        dimensiones=["EI","JP","SN","TF"]
        if len(modelos) == 1:
            print(f"[EJECUCION] Un único modelo detectado. Procesando una sola vez...")
            procesador = Preprocesador(modelos[0])
            procesador.cargar_dataset(nombreCarpeta=carpeta_origen)
            procesador.procesar_dataset(carpeta_dest=carpeta_dest, nom_archivo=f"dataset{dimensiones[0]}")
            
            # Guardamos el dataset 3 veces más con nombres diferentes
            df_procesado = pd.read_parquet(os.path.join(carpeta_dest, f"dataset{dimensiones[0]}.parquet"))
            for dimension in dimensiones[1:]:
                print(f"[EJECUCION] Copiando dataset para dimensión {dimension}...")
                df_procesado.to_parquet(os.path.join(carpeta_dest, f"dataset{dimension}.parquet"), engine="pyarrow")
            
            print("[EJECUCION] Preprocesado completado (una sola vez).")
        else:
            # Si hay múltiples modelos, procesamos cada uno
            print(f"[EJECUCION] Múltiples modelos detectados ({len(modelos)}). Procesando cada uno...")
            for mod, dimension in zip(modelos, dimensiones):
                procesador = Preprocesador(mod)
                procesador.cargar_dataset(nombreCarpeta=carpeta_origen)
                procesador.procesar_dataset(carpeta_dest=carpeta_dest, nom_archivo=f"dataset{dimension}")
                
    balSMOTE = BalanceadorSMOTE(carpeta_dest,balancear=balancear)
    balBORSMOTE = BalanceadorBorderlineSMOTE(carpeta_dest,balancear=balancear)
    balADASYN = BalanceadorADASYN(carpeta_dest,balancear=balancear)
    balENN = BalanceadorENN(carpeta_dest,balancear=balancear)
    balAKNN = BalanceadorAKNN(carpeta_dest,balancear=balancear)
    
    pipelineSMOTE = Pipeline(nombre_modelo=modelo, balanceador=balSMOTE)
    pipelineBORSMOTE = Pipeline(nombre_modelo=modelo, balanceador=balBORSMOTE)
    pipelineADASYN = Pipeline(nombre_modelo=modelo, balanceador=balADASYN)
    pipelineENN = Pipeline(nombre_modelo=modelo, balanceador=balENN)
    pipelineAKNN = Pipeline(nombre_modelo=modelo, balanceador=balAKNN)

    ejecutar_pipelines([pipelineSMOTE, pipelineBORSMOTE, pipelineADASYN, pipelineENN, pipelineAKNN], algoritmo="RL", nombre_archivo=f"Resultados_{modelo.replace('/', '_')}.xlsx", parametros=parametros, buscar_hiperparametros=buscar_hiperparametros)
    ejecutar_pipelines([pipelineSMOTE, pipelineBORSMOTE, pipelineADASYN, pipelineENN, pipelineAKNN], algoritmo="XGBoost", nombre_archivo=f"Resultados_{modelo.replace('/', '_')}.xlsx", parametros=parametros, buscar_hiperparametros=buscar_hiperparametros)
    ejecutar_pipelines([pipelineSMOTE, pipelineBORSMOTE, pipelineADASYN, pipelineENN, pipelineAKNN], algoritmo="LinearSVC", nombre_archivo=f"Resultados_{modelo.replace('/', '_')}.xlsx", parametros=parametros, buscar_hiperparametros=buscar_hiperparametros)
    ejecutar_pipelines([pipelineSMOTE, pipelineBORSMOTE, pipelineADASYN, pipelineENN, pipelineAKNN], algoritmo="MLP", nombre_archivo=f"Resultados_{modelo.replace('/', '_')}.xlsx", parametros=parametros, buscar_hiperparametros=buscar_hiperparametros)
    ejecutar_pipelines([pipelineSMOTE, pipelineBORSMOTE, pipelineADASYN, pipelineENN, pipelineAKNN], algoritmo="KNC", nombre_archivo=f"Resultados_{modelo.replace('/', '_')}.xlsx", parametros=parametros, buscar_hiperparametros=buscar_hiperparametros)
    ejecutar_pipelines([pipelineSMOTE, pipelineBORSMOTE, pipelineADASYN, pipelineENN, pipelineAKNN], algoritmo="DTC", nombre_archivo=f"Resultados_{modelo.replace('/', '_')}.xlsx", parametros=parametros, buscar_hiperparametros=buscar_hiperparametros)
    

def ejecutar_pipelines(pipelines:list[Pipeline], algoritmo:str=None, nombre_archivo="Resultados.xlsx", parametros:dict=None, buscar_hiperparametros:bool=False):
    '''
    Funcion que ejecuta completamente cada Pipeline de la lista pipelines.

    :param pipelines: Lista que contine los diferentes pipelines que se quieren ejecutar
    :param algoritmo: Algoritmo clasificador que se va a usar
    :param nombre_archivo: Nombre del archivo excel donde se quieren guardar los resultados del entrenamiento. Si no existe se creará automaticamente
    :param parametros: Parametros que se quieren usar sobre el Clasificador
    :param buscar_hiperparametros: Bool que indica si quiere que antes del entrenamiento se lleve acabo una busqueda de la mejor combinacion de hiperparametros (Llevará un rato)

    :return: None
    '''
    for pipeline in pipelines:
        pipeline.ejecutar_pipeline_entreno(algotitmo=algoritmo, nombre_archivo=nombre_archivo,parametros=parametros, buscar_hiperparametros=buscar_hiperparametros)

if __name__ == "__main__":
    
    #EJECUCION PIPELINE ROBERTA BASE
    print("="*50)
    print("[INICIO] Ejecución pipeline con Roberta Base ...")
    nombre_modelo = "FacebookAI/roberta-base"
    pipeline_modelo_entreno(nombre_modelo, preprocesar=True, carpeta_origen="dataset9K", modelos=["FacebookAI/roberta-base"], carpeta_dest="datasetRB")
    print("[FIN] Ejecución pipeline con Roberta Base ...")
    print("="*50)
    
    #EJECUCION PIPELINE XLM-ROBERTA-BASE
    print("="*50)
    print("[INICIO] Ejecución pipeline con XLM Roberta Base ...")
    nombre_modelo = "FacebookAI/xlm-roberta-base"
    pipeline_modelo_entreno(nombre_modelo,preprocesar=True, carpeta_origen="dataset9K", modelos=["FacebookAI/xlm-roberta-base"], carpeta_dest="datasetXLMRB")
    print("[FIN] Ejecución pipeline con XLM Roberta Base ...")
    print("="*50)

    #EJECUCION PIPELINE XLM-ROBERTA-LARGE
    print("="*50)
    print("[INICIO] Ejecución pipeline con XLM Roberta Large ...")
    nombre_modelo = "FacebookAI/xlm-roberta-large"
    pipeline_modelo_entreno(nombre_modelo,preprocesar=True, carpeta_origen="dataset9K", modelos=["FacebookAI/xlm-roberta-large"], carpeta_dest="datasetXLMRL")
    print("[FIN] Ejecución pipeline con XLM Roberta Large ...")
    print("="*50)
    
    #EJECUCION PIPELINE ROBERTA BASE FT
    print("="*50)
    print("[INICIO] Ejecución pipeline con Roberta Base FT...")
    modelos = ["./robertaFT/E-I_roberta-base","./robertaFT/J-P_roberta-base","./robertaFT/S-N_roberta-base","./robertaFT/T-F_roberta-base"]
    pipeline_modelo_entreno("Roberta-Base-FT", preprocesar=False, parametros=None, buscar_hiperparametros=False, carpeta_origen="dataset9K", modelos=modelos, carpeta_dest="datasetRBFT")
    print("[FIN] Ejecución pipeline con Roberta Base ...")
    print("="*50)
    
    '''
    '''
