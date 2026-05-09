import os
import joblib
import pandas as pd
import numpy as np
import warnings
from openpyxl import load_workbook
from sklearn.metrics import precision_recall_fscore_support, accuracy_score, confusion_matrix, f1_score
from sklearn.calibration import CalibratedClassifierCV
from xgboost import XGBClassifier
from sklearn.linear_model import LogisticRegression
from sklearn.svm import LinearSVC
from clasificador import MLPC, MBTINet
from skorch import NeuralNetClassifier
import torch.nn as nn
import torch.optim as optim
import torch

# Importaciones de tus módulos locales
from clasificador import XGB, LR, LSVC, MLPC
from balanceador import BalanceadorBorderlineSMOTE, BalanceadorSMOTE, BalanceadorADASYN, BalanceadorSMOTETomek
# Silenciamos warnings para mantener la consola limpia
warnings.filterwarnings("ignore")


def crear_red_gpu(**kwargs):
    device = 'cuda' if torch.cuda.is_available() else 'cpu'
    return NeuralNetClassifier(
        module=MBTINet,
        module__input_dim=768,
        module__hidden_dim=kwargs.get('hidden_dim', 256),
        criterion=nn.CrossEntropyLoss,
        optimizer=optim.Adam,
        lr=kwargs.get('lr', 0.001),
        max_epochs=20, 
        batch_size=kwargs.get('batch_size', 256),
        device=device,
        train_split=None,
        verbose=0
    )

# --- DEFINICIONES GLOBALES ---
RUTA_BASE_MODELOS = "Modelos Definitivos"
DIMENSIONES = ["E-I", "S-N", "T-F", "J-P"]
MODELOS_A_PROBAR = {
    #"XGBoost": (XGB, XGBClassifier),
    #"Regresion_Logistica": (LR, LogisticRegression),
    #"Linear_SVC": (LSVC, LinearSVC),
    "Perceptron_Multicapa": (MLPC, crear_red_gpu)
}


def entrenar_modelos(balanceador):
    '''
    Función original para buscar hiperparámetros, entrenar, calibrar y guardar los modelos.
    (Actualmente no se llama en el main para ahorrar tiempo, pero se conserva por modularidad).
    '''
    print("[INFO] Iniciando Entrenamiento desde cero...")
    if not os.path.exists(RUTA_BASE_MODELOS):
        os.makedirs(RUTA_BASE_MODELOS)
    print(f"[INFO] Modelos entrenados con balanceador '{balanceador.__str__()}' se guardarán en: '{RUTA_BASE_MODELOS+ '/'+balanceador.__str__()}'")
    ruta_balanceador = os.path.join(RUTA_BASE_MODELOS, balanceador.__str__())
    if not os.path.exists(ruta_balanceador):
        os.makedirs(ruta_balanceador)
    
    for nombre_modelo, (WrapperClass, SklearnClass) in MODELOS_A_PROBAR.items():
        print(f"\n" + "="*60)
        print(f" ENTRENANDO FAMILIA: {nombre_modelo.upper()}")
        print("="*60)
        
        wrapper = WrapperClass(balanceador=balanceador)
        
        for dim in DIMENSIONES:
            print(f"\n[+] Entrenando Dimensión: {dim}")
            hiperparametros_optimos = wrapper.busqueda_hiperparametros(dim)
            X_train, y_train = wrapper.datasetEnterno[dim]
            
            modelo_base = SklearnClass(**hiperparametros_optimos)
            modelo_calibrado = CalibratedClassifierCV(
                estimator=modelo_base,
                method="sigmoid",
                cv=4,
                n_jobs=-1
            )
            
            modelo_calibrado.fit(X_train, y_train)
            ruta_final_guardado = os.path.join(ruta_balanceador, f"{dim}_{nombre_modelo}.joblib")
            joblib.dump(modelo_calibrado, ruta_final_guardado)
            print(f"     [OK] Modelo guardado en: {ruta_final_guardado}")


def cargar_modelos(balanceador):
    '''
    Recorre el sistema de archivos, carga todos los modelos pre-entrenados
    y los devuelve estructurados en un diccionario.
    '''
    print(f"\n[INFO] Cargando modelos pre-entrenados desde '{RUTA_BASE_MODELOS}'...")
    modelos_cargados = {}
    
    for nombre_modelo in MODELOS_A_PROBAR.keys():
        modelos_cargados[nombre_modelo] = {}
        for dim in DIMENSIONES:
            ruta_archivo = os.path.join(RUTA_BASE_MODELOS, balanceador.__str__(), f"{dim}_{nombre_modelo}.joblib")
            
            if os.path.exists(ruta_archivo):
                modelo = joblib.load(ruta_archivo)
                modelos_cargados[nombre_modelo][dim] = modelo
                print(f"     [OK] Cargado: {ruta_archivo}")
            else:
                print(f"     [ERROR] Archivo no encontrado: {ruta_archivo}")
                
    return modelos_cargados


def evaluar_y_guardar_formateado(modelos_cargados, balanceador, nombre_excel="Comparativa_Maestra.xlsx", carpeta="resultados"):
    '''
    Coordina la evaluación de todos los modelos cargados, calcula el umbral óptimo para cada uno,
    y guarda los resultados formateados con saltos de línea y filas de parámetros.
    '''
    if not os.path.exists(carpeta):
        os.makedirs(carpeta)
    
    archivo_excel = os.path.join(carpeta, nombre_excel)
    print(f"\n[INFO] Iniciando Evaluación y Exportación Formateada a: {archivo_excel}")

    # Mapeo de datos de validación
    datos_val = {
        "E-I": (np.array(balanceador.val_EI["Embedding"].tolist(), dtype=np.float32), balanceador.val_EI["MBTI"].tolist()),
        "S-N": (np.array(balanceador.val_SN["Embedding"].tolist(), dtype=np.float32), balanceador.val_SN["MBTI"].tolist()),
        "T-F": (np.array(balanceador.val_TF["Embedding"].tolist(), dtype=np.float32), balanceador.val_TF["MBTI"].tolist()),
        "J-P": (np.array(balanceador.val_JP["Embedding"].tolist(), dtype=np.float32), balanceador.val_JP["MBTI"].tolist())
    }

    # Iteramos por cada familia de modelos (XGBoost, LR, etc.)
    for nombre_familia, dict_dimensiones in modelos_cargados.items():
        print(f"\n---> Procesando bloque: {nombre_familia}")
        
        filas_familia = []
        params_familia = ""

        for dim, modelo_calibrado in dict_dimensiones.items():
            X_test, y_test = datos_val[dim]
            
            # --- 1. BÚSQUEDA DE UMBRAL ÓPTIMO (Lógica del script anterior) ---
            proba = modelo_calibrado.predict_proba(X_test)[:, 1]
            mejores_f1 = 0
            umbral_optimo = 0.5
            for umbral in np.arange(0.1, 0.9, 0.01):
                f1_actual = f1_score(y_test, (proba >= umbral).astype(int), average='macro')
                if f1_actual > mejores_f1:
                    mejores_f1 = f1_actual
                    umbral_optimo = umbral
            
            # Predicción final con el mejor umbral
            y_pred = (proba >= umbral_optimo).astype(int)
            
            # --- 2. OBTENCIÓN DE MÉTRICAS FORMATEADAS (Estética de tus funciones) ---
            precision, recall, f1, support = precision_recall_fscore_support(y_test, y_pred, labels=[0, 1])
            acc = accuracy_score(y_test, y_pred)
            cm = confusion_matrix(y_test, y_pred)
            
            # Guardamos los parámetros (solo los extraemos una vez por familia para el Excel)
            if not params_familia:
                params_familia = str(modelo_calibrado.estimator.get_params())

            filas_familia.append({
                "Modelos": f"{nombre_familia} {dim.replace('-', '/')}",
                "Precision": f"0: {precision[0]:.2f}\n1: {precision[1]:.2f}",
                "Recall": f"0: {recall[0]:.2f}\n1: {recall[1]:.2f}",
                "F1-Score": f"0: {f1[0]:.2f}\n1: {f1[1]:.2f}",
                "Support": f"0: {support[0]}\n1: {support[1]}",
                "Accuracy": round(acc, 2),
                "Matriz Confusion": f"[[{cm[0][0]} {cm[0][1]}]\n [{cm[1][0]} {cm[1][1]}]]"
            })

        # --- 3. GUARDADO EN EXCEL (Lógica de overlay y posicionamiento) ---
        df_resultados = pd.DataFrame(filas_familia)
        metodo_balanceo = balanceador.__str__() # Nombre de la pestaña
        
        modo = 'a' if os.path.exists(archivo_excel) else 'w'
        fila_inicio = 0
        escribir_cabecera = True
        
        if modo == 'a':
            try:
                wb = load_workbook(archivo_excel)
                if metodo_balanceo in wb.sheetnames:
                    fila_inicio = wb[metodo_balanceo].max_row + 1
                    escribir_cabecera = False
                wb.close()
            except: 
                pass

        with pd.ExcelWriter(archivo_excel, engine='openpyxl', mode=modo, if_sheet_exists='overlay' if modo == 'a' else None) as writer:
            # Fila de Parámetros
            df_params = pd.DataFrame([["Parametros", params_familia]])
            df_params.to_excel(writer, sheet_name=metodo_balanceo, startrow=fila_inicio, index=False, header=False)
            
            # Tabla de Resultados (2 filas debajo de los parámetros)
            df_resultados.to_excel(writer, sheet_name=metodo_balanceo, startrow=fila_inicio + 2, index=False, header=escribir_cabecera)
            
        print(f"     [OK] Bloque {nombre_familia} guardado.")

    print(f"\n[EXITO] Informe completo generado en: {archivo_excel}")


def main():
    print("[INFO] Arrancando Sistema. Fase: EVALUACIÓN POST-ENTRENAMIENTO")
    
    # 1. Necesitamos el balanceador para obtener los datos de Validación (Test)
    print("[INFO] Preparando datos de validación...")
    balSMOTE = BalanceadorSMOTE(nomCarpeta="datasetRBFT")
    balBorder = BalanceadorBorderlineSMOTE(nomCarpeta="datasetRBFT")
    balADASYN = BalanceadorADASYN(nomCarpeta="datasetRBFT")
    balTomek = BalanceadorSMOTETomek(nomCarpeta="datasetRBFT")
    
    balSMOTE.procesar_todos_ejes()
    balBorder.procesar_todos_ejes()
    balADASYN.procesar_todos_ejes()
    balTomek.procesar_todos_ejes()
    
    #entrenar_modelos(balSMOTE) 
    #entrenar_modelos(balBorder) 
    #entrenar_modelos(balADASYN) 
    #entrenar_modelos(balTomek) 

    # 2. CARGAMOS LOS MODELOS (Nos saltamos el entreno)
    for balanceador in [balSMOTE, balBorder, balADASYN, balTomek]:
        print(f"\n[INFO] Cargando modelos para balanceador: {balanceador.__str__()}")
        modelos_en_memoria = cargar_modelos(balanceador)
        evaluar_y_guardar_formateado(modelos_en_memoria, balanceador)
    



if __name__ == "__main__":
    main()